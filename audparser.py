#!/usr/bin/python3
# -----------------------------------------------------------
# Parser for *.AUD SAP ABAP files. With export to csv or excel.
#
# audparser (C) 2023 Vsevolod Yablonsky
# Released under GNU Public License (GPL)
# email vantela@gmail.com
# -----------------------------------------------------------
#
# Version 1.0
#

import os
import sys
import argparse
import pandas as pd
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl import load_workbook

# Dictionary for type of columns.
cols_dict = {'date': 0, 'time': 1, 'client': 2, 'login': 3, 'terminal': 4, 'tcode': 5, 'report': 6,
				'typecon': 7, 'param': 8, 'eventid': 9, 'osid': 10, 'sapid': 11, 'sapidhex': 12, 'termcut': 13, 'sessionid': 14}
# List for names of columns.
list_of_cols = ["Date", "Time", "Client", "Login", "Terminal", "T-Code", "Report", "TypeConn",
				"Parameters", "EventID", "OSProcID", "SAPProcID", "SAPIDHEX", "termcut", "SessionId"]
# Dictionary for storing information about variables for output.
output_data = {}


def take_block(block):
	"""
	Convert the current string into a list of columns representing one row.
	:param block: list data from file with block_size length
	:return: one-row
	:rtype: list
	"""
	return [
			block[4:8] + '.' + block[8:10] + '.' + block[10:12],        # Date
			block[12:14] + ':' + block[14:16] + ':' + block[16:18],     # Time
			block[112:115],                                             # Client
			block[40:52].strip(),                                       # login
			block[180:].strip(),                                        # Terminal
			block[52:72].strip(),                                       # T-code
			block[72:112].strip(),                                      # Report
			block[30],                                                  # Type of connection (D, B)
			ILLEGAL_CHARACTERS_RE.sub(r'', block[116:180].strip()),     # Variable message data 1&2&3 like param
			block[1:4],                                                 # EventId
			block[18:25],                                               # OS Process ID
			block[25:30],                                               # SAP Process ID
			block[31],                                                  # SAP Process ID in hex
			block[32:40].strip(),                                       # Term-cut
			block[115]                                                  # SessionID
	]


def parsing_for_its_in_args(its, args):
	"""
	Check for pattern in column
	:param its: columns list
	:param args: patterns list
	:return: True if pattern was found in cals, else False
	:rtype: bool
	"""
	if args:
		for arg in args:
			for it in its:
				if arg in it:
					return True
		return False
	else:
		return True


def detect_version(file_name):
	"""
	It reads 2 first bytes from file_name, converts its from hex and match with known strings
	:param file_name: TextIOWrapper
	:return: block size for this file_name
	:rtype: int
	"""
	version = "".join([hex(ord(c))[2:].zfill(2) for c in file_name.read(2)])
	file_name.seek(0)
	match version:
		case '7141':
			return 180  # '4.6c'
		case '3241':
			return 200  # 'non-unicode'
		case '0032' | '3200' | '0478':
			return 400  # 'unicode'
		case _:
			print("Failed to detect block size: ", version)
			sys.exit()


def remove_extra_cols(element):
	"""
	Remove extra columns from element
	:param element: one row (list of columns)
	:return: list   without columns from parsed_args.remove
	:rtype: list
	"""
	if parsed_args.remove:
		cols_for_remove = []
		for col in parsed_args.remove:
			cols_for_remove.append(cols_dict[col])
		cols_for_remove.sort(reverse=True)
		for col in cols_for_remove:
			del element[col]
	return element


def parse_file(filename):
	"""
	Open file and read it
	:param filename: the next file name for reading
	:return: list of columns
	:rtype: list
	"""
	res = []
	with open(filename, encoding='latin-1') as f:
		block_size = detect_version(f)
		block = f.read(block_size)
		# We don't have non-unicode systems or 4.6c. Next code only for block_size = 400 and 200 symbols
		while block != '':
			item = take_block(block[0:len(block):2])
			if all([parsing_for_its_in_args([item[cols_dict['typecon']]], parsed_args.typecon),
					parsing_for_its_in_args([item[cols_dict['terminal']]], parsed_args.terminal),
					parsing_for_its_in_args([item[cols_dict['login']]], parsed_args.login),
					parsing_for_its_in_args([item[cols_dict['tcode']], item[cols_dict['param']]], parsed_args.tcode),
					parsing_for_its_in_args([item[cols_dict['report']]], parsed_args.report),
					parsing_for_its_in_args([item[cols_dict['client']]], parsed_args.client)]):
				res.append(remove_extra_cols(item))
			block = f.read(block_size)
		f.close()
	return res


def print_results(res):
	"""
	Print on display
	:param res: list of columns
	"""
	for row in res:
		for col in row:
			print(f"{col}", end="\t")
		print()


def csv_export(res):
	"""
	Write columns to output_data['cvs_export_file'] separated by semicolons
	:param res: list of columns
	"""
	for row in res:
		for col in row:
			output_data['cvs_export_file'].write(f"{col};")
		output_data['cvs_export_file'].write("\n")


def excel_export(res):
	"""
	Write columns to output_data['xlsx_export_file'] Excel file.
	Only 1 000 000 rows on sheet. Create a new sheet for the following rows..
	:param res: list of columns
	"""
	rows_per_sheet = 1_000_000
	if 'sheet' not in output_data.keys():
		output_data['sheet'] = 0
	if 'start_row' not in output_data.keys():
		output_data['start_row'] = 0
	start_row_init = output_data['start_row']
	start_index = 0
	end_index = rows_per_sheet - output_data['start_row']
	while start_index <= len(res):
		df = pd.DataFrame(list(res[start_index:end_index]))
		df.to_excel(output_data['xlsx_export_file'], index=False, startrow=output_data['start_row'], header=False,
					sheet_name='sheet_' + str(output_data['sheet']))
		start_index = end_index
		end_index = end_index + rows_per_sheet
		output_data['start_row'] = 0
		output_data['sheet'] = output_data['sheet'] + 1
	# Save current sheet number and start_row for the next part of data
	output_data['sheet'] = output_data['sheet'] - 1
	output_data['start_row'] = (start_row_init + len(res)) % rows_per_sheet


def export_data(res):
	"""
	Prepare some variables and run functions for print/cvs/excel output
	:param res: list of columns
	"""
	if parsed_args.header and ('header' not in output_data.keys()):
		# add header if it needs
		res.insert(0, remove_extra_cols(list_of_cols))
		output_data['header'] = True
	
	if parsed_args.csv:
		if 'cvs_export_file' not in output_data.keys():
			output_data['cvs_export_file'] = open(f'{parsed_args.export_name}.csv', parsed_args.overwrite)
		csv_export(res)
		
	if parsed_args.excel:
		if 'xlsx_export_file' not in output_data.keys():
			if os.path.isfile(f'{parsed_args.export_name}.xlsx') and parsed_args.overwrite == 'a':
				output_data['xlsx_export_file'] = pd.ExcelWriter(f'{parsed_args.export_name}.xlsx', engine='openpyxl',
																mode='a', if_sheet_exists='overlay')
				output_data['workbook'] = load_workbook(f'{parsed_args.export_name}.xlsx')
				sheet_names = output_data['workbook'].sheetnames
				output_data['sheet'] = len(sheet_names) - 1
				output_data['start_row'] = output_data['workbook'][sheet_names[-1]].max_row
			else:
				output_data['xlsx_export_file'] = pd.ExcelWriter(f'{parsed_args.export_name}.xlsx', engine='openpyxl')
		excel_export(res)
		
	if parsed_args.print:
		print_results(res)


def main():
	input_files = set()
	# prepare file list for parsing
	for name in parsed_args.aud:
		if os.path.isfile(name):
			input_files.add(os.path.abspath(name))
		elif os.path.isdir(name):
			for path, dirs, files in os.walk(name):
				for file in files:
					current_file = os.path.join(path, file)
					if '.AUD' in current_file:
						input_files.add(os.path.abspath(current_file))
	# Print main information about this run
	print("We start the parsing with: ", input_files)
	if parsed_args.remove:
		print("Cols for remove: ", parsed_args.remove)
	if parsed_args.terminal:
		print("Filter by terminal(in fields terminal and fqdn): ", parsed_args.terminal)
	if parsed_args.login:
		print("Filter by login: ", parsed_args.login)
	if parsed_args.tcode:
		print("Filter by T-Code(in fields tcode and param): ", parsed_args.tcode)
	if parsed_args.report:
		print("Filter by report: ", parsed_args.report)
	if parsed_args.client:
		print("Filter by client: ", parsed_args.client)
	if parsed_args.typecon:
		print("Filter by type connection: ", parsed_args.typecon)
	
	# main loop for parse all files
	collected_rows = 0
	for file_name in input_files:
		result = parse_file(file_name)
		collected_rows = collected_rows + len(result)
		export_data(result)
	
	print("Collected rows:", collected_rows)
	# close export files if it needs and hints about export parameters
	if parsed_args.csv:
		output_data['cvs_export_file'].close()
	else:
		print("For exporting to csv plz use -csv option")
	if parsed_args.excel:
		output_data['xlsx_export_file'].close()
	else:
		print("For exporting to excel plz use -excel option")
	if not parsed_args.print:
		print("For printing on display plz use -print option")
	

if __name__ == '__main__':
	parser = argparse.ArgumentParser()
	parser.add_argument('-remove', metavar='eventid osid sapid sapidhex termcut sessionid', nargs='*',
						help='You can specify cols for remove. If you use this option, default removing will be '
							'overwritten. Other cols: date time client login terminal tcode report typecon param',
						default=['eventid', 'osid', 'sapid', 'sapidhex', 'termcut', 'sessionid'])
	parser.add_argument('-terminal', metavar='31709', nargs='*',
						help='It tries to search this in fields terminal and fqdn, as substring')
	parser.add_argument('-login', metavar='SAPROOT CUASM7', nargs='*',
						help='It tries to search this in field login, as substring')
	parser.add_argument('-tcode', metavar='SU01 PFCG', nargs='*',
						help='It tries to search this in fields tcode and param, as substring')
	parser.add_argument('-report', metavar='ZBC', nargs='*',
						help='It tries to search this in field report, as substring')
	parser.add_argument('-client', metavar='000 300', nargs='*',
						help='It tries to search this in field client, as substring')
	parser.add_argument('-typecon', metavar='D B', nargs='*',
						help='It tries to search this in field typecon, as substring')
	parser.add_argument('-header', help='Add header for cols', action='store_true')
	parser.add_argument('-print', help='Print all', action='store_true')
	parser.add_argument('-excel', help='Enable export to excel with default name results.xlsx', action='store_true')
	parser.add_argument('-csv', help='Enable export to csv with default name results.csv', action='store_true')
	parser.add_argument('-export_name', metavar='results', help='use this name for file', default="results")
	parser.add_argument('-overwrite', help='Overwrite existing files with parsing results, default: append',
						action='store_true')
	parser.add_argument('-aud', nargs='*', help='parse all *.AUD from this directory or file, "./" by default',
						default=".")
	parsed_args = parser.parse_args()
	parsed_args.overwrite = 'w' if parsed_args.overwrite else 'a'
	main()
